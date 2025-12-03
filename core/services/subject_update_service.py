import sys
from tkinter import messagebox

from openpyxl import load_workbook
import os
import re
from collections import defaultdict
from typing import Any
import time
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from copy import copy
from openpyxl.styles import PatternFill
import pandas as pd

class SubjectUpdateService:
    """
    科目更新邏輯：
    - 從 Excel 讀取「科目表」與「分類帳」
    - 根據代號比對，更新本月新增明細
    - 標示異動列
    """

    def __init__(self, file_path: str, logger=None, app=None):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"找不到檔案：{file_path}")

        self.file_path = file_path
        self.wb_values = load_workbook(file_path, data_only=True)
        self.wb = load_workbook(file_path, data_only=False)

        self.logger = logger or (lambda msg: print(msg))
        self.app = app  # ExcelToolApp 實例（可為 None）

        # 紀錄分類帳中「含非法符號」的科目名稱
        self.invalid_items = []

    def _check_cancel(self):
        """隨時可以在迴圈裡呼叫，一旦使用者按了停止就丟 Exception 中斷流程"""
        if self.app is not None and getattr(self.app, "cancel_requested", False):
            # 這個訊息會被上層捕捉並顯示
            raise RuntimeError("使用者已中止科目更新作業。")

    def _log(self, msg: str):
        """統一 logging 介面"""
        self.logger(msg)

    # ---------------------------------------------------------
    # 🔍 找出「分類帳」分頁
    # ---------------------------------------------------------
    def find_ledger_sheet(self) -> str:
        """模糊搜尋「分類帳」分頁名稱（忽略全形／半形空白）"""
        for name in self.wb_values.sheetnames:
            normalized = name.replace(" ", "").replace("　", "")
            if "分類帳" in normalized:
                return name
        available = "、".join(self.wb_values.sheetnames)
        raise ValueError(f"❌ 找不到『分類帳』工作表（目前分頁：{available}）")

    # ---------------------------------------------------------
    # 🧭 主函式
    # ---------------------------------------------------------
    def check_subject_sheet_existence(self, target_month: str):
        """主函式：綜合執行三個子步驟"""
        # 先清空上一輪的紀錄
        self.invalid_items = []
        sheet = self.wb_values[self.find_ledger_sheet()]
        self._check_cancel()  # ⭐ 加這行
        rows = self._filter_valid_rows(sheet, target_month)
        self._check_cancel()  # ⭐ 加這行

        # ★ 如果有非法字元的科目名稱，直接在這裡用 _compose_message 擋掉
        if self.invalid_items:
            return self._compose_message(
                zero_items_but_kept=[],
                inconsistent=[],  # 餘額不符先給空
                target_month=target_month,
                invalid_items=self.invalid_items
            )

        latest_rows, zero_items_but_kept = self._get_last_rows_by_item(rows)
        self._check_cancel()  # ⭐ 加這行
        return self._check_sheet_existence_and_print(latest_rows, target_month, zero_items_but_kept)

    # ---------------------------------------------------------
    # 🧩 Step 1️⃣ 篩出符合條件的列
    # ---------------------------------------------------------
    def _filter_valid_rows(self, sheet, target_month: str):
        """篩出所有符合條件的列"""
        month_extract = re.compile(r"^(1\d{2})[-/.]?(0[1-9]|1[0-2])")
        target_int = int(target_month)
        valid_rows = []

        for row in sheet.iter_rows(min_row=2):
            if not self._validate_row(row, target_int, month_extract, debug=False):
                continue

            a_val = str(row[0].value).strip() if row[0].value else ""
            d_val_raw = str(row[3].value).strip() if row[3].value else ""
            c_val = str(row[2].value).strip() if row[2].value else ""
            row_number = row[0].row

            # 🔴 D 欄科目名稱若含非法字元 → 記錄起來，不讓它進入後續流程
            if any(ch in d_val_raw for ch in self.INVALID_SHEET_CHARS):
                # 紀錄成「第X列：名稱」這種可讀格式
                self.invalid_items.append(f"第{row_number}列：{d_val_raw}")
                continue

            d_val = d_val_raw
            i_val = float(row[8].value)

            valid_rows.append((row_number, a_val, d_val, i_val,c_val))

        return valid_rows

    # ---------------------------------------------------------
    # 🧩 Step 2️⃣ 同一項目取最後一筆
    # ---------------------------------------------------------
    def _get_last_rows_by_item(self, valid_rows):
        """
        對同一項目保留最後一筆（行號最大者）；
        若該項目的最後一筆 I 欄為 0：
          - 若該項目不存在於工作表 → 排除
          - 若該項目存在於工作表 → 保留並標記
        """
        rows_by_item = defaultdict(list)
        for row_number, a_val, d_val, i_val ,c_val in valid_rows:
            rows_by_item[d_val].append((row_number, a_val, i_val,c_val))

        latest_rows = {}
        zero_items_but_kept = []

        for d_val, rows in rows_by_item.items():
            self._check_cancel()  # ⭐ 加這行
            rows.sort(key=lambda x: x[0])
            row_number, a_val, i_val ,c_val= rows[-1]

            if i_val is None or float(i_val) == 0:
                if not self._check_item_in_sheet(d_val):
                    self._log(f"🗑️ 項目【{d_val}】最後餘額為 0 且無對應工作表，已排除。")
                    continue
                else:
                    zero_items_but_kept.append(d_val)
                    self._log(f"⚠️ 項目【{d_val}】最後餘額為 0，但仍存在於工作表，已保留。")

            latest_rows[d_val] = (row_number, a_val, i_val, c_val)

        return latest_rows, zero_items_but_kept

    def _get_active_items(self, valid_rows):
        """
        傳回區間內所有「有明細」的科目名稱（d_val）
        valid_rows = [(row_num, a_val, d_val, i_val), ...]
        """
        items = set()
        for row_number, a_val, d_val, i_val in valid_rows:
            d = d_val.strip() if isinstance(d_val, str) else None
            if d:
                items.add(d)
        return items

    def _check_item_in_sheet(self, item_code: str) -> bool:
        """檢查指定項目代號是否存在於工作表中。"""
        sheetnames = [
            s.title.replace(" ", "").replace("　", "")
            for s in self.wb_values.worksheets if s.sheet_state == "visible"
        ]
        clean_name = item_code.replace(" ", "").replace("　", "")
        return clean_name in sheetnames

    # ---------------------------------------------------------
    # 🧩 Step 3️⃣ 驗證單列
    # ---------------------------------------------------------
    def _validate_row(self, row, target_int: int, month_extract, debug=False) -> bool:
        """檢查單列是否符合條件"""
        row_num = row[0].row
        a_val = str(row[0].value).strip() if row[0].value else ""
        c_val = str(row[2].value).strip() if row[2].value else ""
        d_val = str(row[3].value).strip() if row[3].value else ""
        i_val = row[8].value

        m = month_extract.match(a_val)
        if m:
            a_month_int = int(m.group(1) + m.group(2))
            if a_month_int > target_int:
                # 超過 target 月份，可能需要特例
                if a_val != "上期結轉":
                    return False
            # 日期合法，直接通過
        else:
            # 無法解析日期 → 只有是「上期結轉」才通過
            if a_val != "上期結轉":
                return False
            a_month_int = target_int  # 可以視為最新月份
        try:
            float(i_val)
        except (TypeError, ValueError):
            return False
        if not c_val or c_val[0] not in ("1", "2"):
            return False
        if not d_val:
            return False
        return True

    # ---------------------------------------------------------
    # 🧩 Step 4️⃣ 餘額比對
    # ---------------------------------------------------------
    def _compare_balance(self, ws, ledger_i, target_month):
        """比對工作表中的最後一筆 I 欄餘額（A、C、D、I 欄不可為NONE）"""
        matched_rows = []

        for row in ws.iter_rows(min_row=2):

            a_val = str(row[0].value).strip() if row[0].value else ""
            c_val = str(row[2].value).strip() if row[2].value else ""
            d_val = str(row[3].value).strip() if row[3].value else ""
            i_val = row[8].value

            # A、C、D 欄必須有值（空字串或 None 都算空）
            if a_val is None or str(a_val).strip() == "":
                continue
            if c_val is None or str(c_val).strip() == "":
                continue
            if d_val is None or str(d_val).strip() == "":
                continue

            # I欄必須有值，可以是 0，但不能是 None 或空字串
            if i_val is None or str(i_val).strip() == "":
                continue

            matched_rows.append((row[0].row, float(i_val)))

        if not matched_rows:
            return None, None, False

        sheet_row, sheet_i = matched_rows[-1]
        same = abs(ledger_i - sheet_i) < 0.001
        return sheet_row, sheet_i, same

    # 🔸 統一管理 Excel 禁用的工作表字元
    INVALID_SHEET_CHARS = (":", "\\", "/", "?", "*", "[", "]")

    # ---------------------------------------------------------
    # 🧩 Step 5️⃣ 組合訊息
    # ---------------------------------------------------------
    def _compose_message(self, zero_items_but_kept, inconsistent, target_month, invalid_items=None):
        """組合最終訊息，避免重複並輸出清楚分類"""
        invalid_items = invalid_items or []
        inconsistent = [x for x in inconsistent if x not in zero_items_but_kept]
        parts = []

        status = "success"
        # ✅ status 也要把 invalid_items 納入
        if zero_items_but_kept or inconsistent or invalid_items:
            status = "error"

        # 🔴 先處理「名稱含非法字元」的情況
        if invalid_items:
            parts.append(
                "⚠️ 以下會計項目名稱包含 Excel 不允許的符號 "
                f"({', '.join(self.INVALID_SHEET_CHARS)})，請修改科目名稱後再重新執行：\n  "
                + "、".join(sorted(set(invalid_items)))
            )

        if zero_items_but_kept:
            parts.append(
                f"⚠️ 文中系統目前 {target_month} 月以下會計項目非為 0，且未有科餘分頁，"
                f"請確認分頁名稱及分頁內容後再重新執行：\n  " + "、".join(zero_items_but_kept)
            )

        if inconsistent:
            parts.append(
                f"⚠️ 以下會計項目之分頁餘額與文中系統目前 {target_month} 月餘額不符，"
                f"請先確認餘額數或分頁名稱後再重新執行：\n  " + "、".join(list(dict.fromkeys(inconsistent)))
            )

        # ✅ 全部都沒問題，才印出 ✅
        if not zero_items_but_kept and not inconsistent and not invalid_items:
            parts.append(f"✅ 所有項目均與文中系統 {target_month} 月餘額一致。")

        msg = "\n\n".join(parts)
        self._log(msg)
        return {
            "status": status,
            "message": msg,
            "details": {
                "inconsistent": inconsistent,
                "zero_items_but_kept": zero_items_but_kept,
                "invalid_items": invalid_items,  # ✅ 多回傳這個
            }
        }

        # ---------------------------------------------------------
        # 🧩 Step 6️⃣ 主比對函式 (修正版)
        # ---------------------------------------------------------

    def _check_sheet_existence_and_print(self, latest_rows, target_month, zero_items_but_kept=None):
        """比對分頁是否存在並印出結果"""
        inconsistent = []

        # 🔴 修改 1：建立 { '去空白名稱': '真正的分頁名稱' } 的對照表
        # 這樣就算分頁名稱有多餘空白，我們也能透過乾淨的名稱找到它真正的 Key
        clean_to_real_map = {
            s.title.replace(" ", "").replace("　", ""): s.title
            for s in self.wb_values.worksheets if s.sheet_state == "visible"
        }
        # 🔴【排除清單】這五個代號將被跳過餘額比對
        EXCLUDED_CODES = ["1191", "1192", "1193", "1197", "1198"]
        if not latest_rows:
            msg = f"❌ 沒有符合條件的資料（<= {target_month}）"
            self._log(msg)
            return {"status": "error", "message": msg, "details": {}}

        for d_val, (ledger_row, ledger_date, ledger_i, ledger_c) in  sorted(latest_rows.items(), key=lambda x: int(x[1][3])) :

            # 🔴【執行排除】檢查代號是否在排除清單內
            if ledger_c in EXCLUDED_CODES:
                self._log(f"ℹ️ 科目代號【{ledger_c}】已設定為排除，跳過餘額比對。")
                continue
            # 這是分類帳上的科目名稱（已去除前後空白，但中間可能有空白）
            clean_name = d_val.replace(" ", "").replace("　", "")

            # 🔴 修改 2：改查對照表，而不是查 list
            if clean_name not in clean_to_real_map:
                inconsistent.append(d_val)
                continue

            # 🔴 修改 3：取得「真正的分頁名稱」來開啟 Worksheet
            real_sheet_name = clean_to_real_map[clean_name]

            try:
                ws = self.wb_values[real_sheet_name]
            except KeyError:
                # 雙重保險：理論上不會發生，但如果發生了就視為找不到
                inconsistent.append(d_val)
                continue

            sheet_row, sheet_i, same = self._compare_balance(ws, ledger_i, target_month)
            if sheet_row is None or not same:
                inconsistent.append(d_val)

        return self._compose_message(zero_items_but_kept, inconsistent, target_month)

    # ---------------------------------------------------------
    # 🧭 外部呼叫介面
    # ---------------------------------------------------------
    def run_check(self, latest_month) -> dict:
        """
        執行完整檢查：
        - 若有錯誤：回傳 status="error"
        - 若一致：回傳 status="success"
        """
        result = self.check_subject_sheet_existence(latest_month)
        return result

    # ----------------------------------------------------------------
    def update_subject_sheets(self, make_month: str, latest_month: str):
        """
        依據資產負債表與分類帳，自動更新各科目分頁
        make_month: 製作科餘年月（例：11408）
        latest_month: 最新科餘年月（例：11406）
        """
        self._log(f"🧭 開始更新科目分頁：製作科餘月={make_month}，最新科餘月={latest_month}")

        # 1️⃣ 找出資產負債表工作表
        balance_sheet = self.wb_values["資產負債表"]

        # 2️⃣ 掃描 A、D 欄，找出代號與對應名稱
        subject_map = self._extract_subjects_from_balance(balance_sheet)

        # 3️⃣ 從分類帳取得要複製的資料列
        ledger_sheet = self.wb_values[self.find_ledger_sheet()]
        records_to_copy = self._find_records_in_ledger(ledger_sheet, subject_map, make_month, latest_month)

        # 4️⃣ 寫入對應的科目分頁
        self._insert_records_into_sheets(records_to_copy, make_month, latest_month)

        msg = "✅ 科目更新完成。"
        self._log(msg)
        self._popup(msg)

    def _extract_subjects_from_balance(self, sheet):
        """從資產負債表中抓出項目代號與名稱"""
        subjects = {}
        # 🔴【新增】排除代號清單 (這是資產負債表端篩選)
        EXCLUDED_CODES = ["1191", "1192", "1193", "1197", "1198"]
        def clean(s):
            if not s:
                return ""

            text = str(s)

            # ① 移除所有空白（含半形、全形）
            text = "".join(text.split())

            # ② 移除「減：」開頭（含全形/半形/變種冒號）
            # 常見符號：: ： ︰ ﹕ ｡ ．
            # 支援繁簡體「減 / 减」
            remove_prefixes = [
                "減:", "減：", "減︰", "減﹕",
                "减:", "减：", "减︰", "减﹕",
            ]

            for prefix in remove_prefixes:
                if text.startswith(prefix):
                    text = text[len(prefix):]
                    break  # 移除一次即可，避免重複

            return text
        # 可能要改
        for row in sheet.iter_rows(min_row=2):
            a_val = clean(row[0].value)
            b_val = clean(row[1].value)
            d_val = clean(row[3].value)
            e_val = clean(row[4].value)

            # 🔴 修正：檢查代號是否在排除清單內
            if a_val.startswith(("1", "2")) and b_val:
                if a_val not in EXCLUDED_CODES:
                    subjects[b_val] = a_val

            # 🔴 修正：檢查代號是否在排除清單內
            if d_val.startswith(("1", "2")) and e_val:
                if d_val not in EXCLUDED_CODES:
                    subjects[e_val] = d_val

        self._log(f"📘 共找到 {len(subjects)} 個項目：{list(subjects.values())[:5]}...")
        return subjects

    def _find_records_in_ledger(self, sheet, subject_map, make_month, latest_month):
        """找出分類帳中介於最新科餘月 ~ 製作月的明細"""
        records = []
        start_int = int(latest_month)
        end_int = int(make_month)

        for row in sheet.iter_rows(min_row=2):
            a_val = str(row[0].value).strip() if row[0].value else ""
            d_val = str(row[3].value).strip() if row[3].value else ""
            if not a_val or not d_val:
                continue

            # 轉換日期 114-08-05 → 11408
            m = re.match(r"^(1\d{2})[-/.]?(0[1-9]|1[0-2])", a_val)
            if not m:
                continue
            month_int = int(m.group(1) + m.group(2))
            if not (start_int < month_int <= end_int):
                continue
            #
            if d_val in subject_map:
                records.append((d_val, row[:9]))
        self._log(f"找到要貼入的紀錄：{[(d_val, [c.value for c in row[:9]]) for d_val, row in records]}")

        self._log(f"📗 找到 {len(records)} 筆新資料。")
        return records

    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    def _insert_records_into_sheets(self, records, make_month, latest_month):
        """將分類帳的新資料寫入各自的科目分頁，若無則建立"""

        ledger_ws_src = self.wb["分類帳"]
        updated_sheets = set()  # ← 新增：記錄本次有更新的分頁名稱

        for subject_code, row_cells in records:

            # ------ 判斷工作表名稱 ------
            # 先去掉空白比對
            clean_subject = subject_code.replace(" ", "").replace("　", "")

            # 先找是否有隱藏的同名分頁
            hidden_sheets = {s.title.replace(" ", "").replace("　", ""): s for s in self.wb.worksheets if
                             s.sheet_state == "hidden"}
            visible_sheets = {s.title.replace(" ", "").replace("　", ""): s for s in self.wb.worksheets if
                              s.sheet_state == "visible"}

            if clean_subject in visible_sheets:
                # 已存在可見分頁，直接使用
                ws = visible_sheets[clean_subject]
            elif clean_subject in hidden_sheets:
                # 已存在隱藏分頁，加 @ 後建立新的分頁
                new_name = f"@{subject_code}"
                if new_name in self.wb.sheetnames:
                    ws = self.wb[new_name]  # 已有 @ 分頁，直接使用
                else:
                    ws = self.wb.create_sheet(new_name)
                    # 複製欄寬與標頭列
                    for col in ledger_ws_src.column_dimensions:
                        ws.column_dimensions[col].width = ledger_ws_src.column_dimensions[col].width
                    for col_idx, cell in enumerate(ledger_ws_src[1], start=1):
                        new_cell = ws.cell(row=1, column=col_idx, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                    self._log(f"🆕 建立新隱藏分頁並複製完整標頭：{new_name}")
            else:
                # 完全不存在，直接建立原名分頁
                ws = self.wb.create_sheet(subject_code)
                for col in ledger_ws_src.column_dimensions:
                    ws.column_dimensions[col].width = ledger_ws_src.column_dimensions[col].width
                for col_idx, cell in enumerate(ledger_ws_src[1], start=1):
                    new_cell = ws.cell(row=1, column=col_idx, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
                self._log(f"🆕 建立新工作表並複製完整標頭：{subject_code}")

            # ------ 找最後一列 ------
            last_row = 1
            for r in range(2, ws.max_row + 1):
                a_val = ws[f"A{r}"].value
                c_val = ws[f"C{r}"].value
                d_val = ws[f"D{r}"].value
                i_val = ws[f"I{r}"].value

                # 檢查 A, C, D 欄位：必須有內容且不是空字串/空白
                a_is_valid = a_val is not None and str(a_val).strip() != ""
                c_is_valid = c_val is not None and str(c_val).strip() != ""
                d_is_valid = d_val is not None and str(d_val).strip() != ""

                # 檢查 I 欄位：必須有值 (is not None)，數值可以是 0 (因為 0 is not None 是 True)
                i_is_valid = i_val is not None

                if a_is_valid and c_is_valid and d_is_valid and i_is_valid:
                    last_row = r

            # ------ 插入新資料 ------
            insert_row = last_row + 1
            ws.insert_rows(insert_row)

            for col_idx, src_cell in enumerate(row_cells, start=1):
                dest = ws.cell(row=insert_row, column=col_idx, value=src_cell.value)
                if src_cell.has_style:
                    dest.font = copy(src_cell.font)
                    dest.border = copy(src_cell.border)
                    dest.fill = copy(src_cell.fill)
                    dest.number_format = copy(src_cell.number_format)
                    dest.protection = copy(src_cell.protection)
                    dest.alignment = copy(src_cell.alignment)

            self._mark_sheet_colors(ws)
            updated_sheets.add(subject_code)
            self._log(f"📄 已插入 {subject_code} 第 {insert_row} 列")

        # ----------------------------------------------------
        # 🔹 呼叫獨立方法建立更新清單工作表
        # ----------------------------------------------------
        self._create_update_summary_sheet(updated_sheets, make_month, latest_month)
        # ------ 儲存 ------
        base, ext = os.path.splitext(self.file_path)
        new_path = base + "_updated" + ext

        self.wb.save(new_path)
        self._log(f"💾 已另存新檔：{new_path}")
        # # ------ 儲存 ------
        # self.wb.save(self.file_path)
        # self._log(f"💾 已儲存更新結果。")

    # 分頁操作紀錄
    def _create_update_summary_sheet(self, updated_sheets, make_month, latest_month):
        """建立本次更新清單工作表，並設為隱藏。"""

        if not updated_sheets:
            self._log(f"ℹ️ 本次沒有任何分頁被更新。")
            return

        summary_sheet_name = f"更新清單_{make_month}"

        # 若已存在同名分頁 → 先刪除
        if summary_sheet_name in self.wb.sheetnames:
            del self.wb[summary_sheet_name]

        ws_summary = self.wb.create_sheet(summary_sheet_name)

        ws_summary["A1"] = "科目代號（分頁名稱）"
        ws_summary["B1"] = "製作科餘月"
        ws_summary["C1"] = "最新科餘月"

        for idx, name in enumerate(sorted(updated_sheets), start=2):
            ws_summary[f"A{idx}"] = name
            ws_summary[f"B{idx}"] = make_month
            ws_summary[f"C{idx}"] = latest_month

        # 設為隱藏
        ws_summary.sheet_state = "hidden"

        self._log(f"📝 已建立本次更新清單工作表：{summary_sheet_name}")

    def _mark_sheet_colors(self, ws):
        from openpyxl.styles import PatternFill

        yellow_fill = PatternFill(start_color="FFF6D6A8", end_color="FFF6D6A8", fill_type="solid")
        red_fill = PatternFill(start_color="FFE1E5E9", end_color="FFE1E5E9", fill_type="solid")

        # ------------------------
        # 第 1 步：一次把 E/F/G 欄資料收集起來
        # ------------------------
        E_vals = {}
        F_vals = []
        G_vals = []

        for r in range(2, ws.max_row + 1):
            E_vals.setdefault(ws[f"E{r}"].value, []).append(r)

            F_vals.append((r, ws[f"F{r}"].value))
            G_vals.append((r, ws[f"G{r}"].value))

        # E 欄重複的值
        duplicated_E = {k: rows for k, rows in E_vals.items() if k and len(rows) >= 2}

        # 把所有 F/G 數值變成 set（便於比對）
        F_set = set()
        G_set = set()
        for _, v in F_vals:
            try:
                F_set.add(float(v))
            except:
                pass
        for _, v in G_vals:
            try:
                G_set.add(float(v))
            except:
                pass

        # ------------------------
        # 第 2 步：開始標色
        # ------------------------

        # 規則 1：E 欄文字重複 → 黃色
        for rows in duplicated_E.values():
            for r in rows:
                ws[f"E{r}"].fill = yellow_fill

        # 規則 2：任一列 F == 任一列 G → 紅色
        for r, v in F_vals:
            try:
                fv = float(v)
                if fv != 0 and fv in G_set:  # 新增 fv != 0
                    ws[f"F{r}"].fill = red_fill
                    ws[f"G{r}"].fill = red_fill
            except:
                pass

        for r, v in G_vals:
            try:
                gv = float(v)
                if gv != 0 and gv in F_set:  # 新增 gv != 0
                    ws[f"F{r}"].fill = red_fill
                    ws[f"G{r}"].fill = red_fill
            except:
                pass

    def run_copy_data(self, make_month, latest_month):
        """執行檢查通過後的下一步：更新科目分頁"""
        self.update_subject_sheets(make_month, latest_month)

    def _popup(self, msg: str):
        """讓 Service 可以安全叫出彈窗（需要 app 才能 after 回主執行緒）"""
        if hasattr(self, "app") and self.app:
            self.app.after(0, lambda: messagebox.showinfo("完成", msg))
