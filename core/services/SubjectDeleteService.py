# core/services/subject_delete_service.py
from openpyxl import load_workbook
from collections import defaultdict
import os


class SubjectDeleteService:
    """
    科目明細刪除邏輯（第四模組）

    流程：
    1. 開啟檔案與「更新清單_XXXX」工作表
    2. 確認更新清單 B/C 欄與目前輸入的年月一致
    3. 依更新清單中的科目代號，逐一處理各科目分頁：
       - 若工作表不存在 → 記錄 log，略過
       - 以「摘要（E 欄）」分組，分別加總 F 欄與 G 欄
       - 若某個摘要下 F、G 加總相等 → 刪除該摘要的所有列
    """

    def __init__(self, file_path: str, logger=None, app=None):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"找不到檔案：{file_path}")

        self.file_path = file_path
        # 刪除需要真正的活體 workbook
        self.wb = load_workbook(file_path, data_only=False)

        # logger：預設印到 console；若從 GUI 進來會是 app.append_log
        self.logger = logger or (lambda msg: print(msg))
        # app：用來支援「立即停止執行」的 cancel flag（可為 None）
        self.app = app

    # ---------- 共用工具 ----------

    def _log(self, msg: str):
        self.logger(msg)

    def _check_cancel(self):
        """若 GUI 設定了 cancel_requested，就中止整個流程。"""
        if self.app is not None and getattr(self.app, "cancel_requested", False):
            raise RuntimeError("使用者已中止科目明細刪除作業。")

    # ---------- 主流程 ----------

    def run_delete(self, make_month: str, latest_month: str) -> str:
        """
        對指定的「製作科餘月 / 最新科餘月」執行刪除流程。
        回傳一段訊息（給狀態列或彈窗用）
        """
        summary_name = f"更新清單_{make_month}"

        if summary_name not in self.wb.sheetnames:
            raise Exception(
                f"找不到更新清單工作表「{summary_name}」，"
                f"請先執行『科目更新』工具（第三步驟）。"
            )

        ws_summary = self.wb[summary_name]
        self._log(f"📄 使用更新清單工作表：{summary_name}")

        # 1️⃣ 讀取更新清單 + 檢查 B/C 是否符合目前輸入的年月
        subjects = self._load_and_validate_summary(ws_summary, make_month, latest_month)

        if not subjects:
            msg = "ℹ️ 更新清單中沒有任何科目代號可供刪除。"
            self._log(msg)
            return msg

        # 2️⃣ 逐一處理各科目分頁
        total_deleted_rows = 0
        processed_sheets = 0

        for subject_code in subjects:
            self._check_cancel()

            deleted = self._process_subject_sheet(subject_code)
            if deleted is None:
                # 分頁不存在 → 已在內部 log，略過
                continue

            processed_sheets += 1
            total_deleted_rows += deleted
            self._log(f"🧹 分頁「{subject_code}」刪除 {deleted} 列。")

        # 3️⃣ 儲存結果
        self.wb.save(self.file_path)
        self._log("💾 刪除結果已儲存。")

        summary_msg = (
            f"✅ 科目明細刪除完成。共處理 {processed_sheets} 個分頁，"
            f"刪除 {total_deleted_rows} 列。"
        )
        self._log(summary_msg)
        return summary_msg

    # ---------- Step 1：讀取 & 驗證更新清單 ----------

    def _load_and_validate_summary(self, ws_summary, make_month: str, latest_month: str):
        """
        讀取更新清單：
        - A 欄：科目代號（分頁名稱）
        - B 欄：製作科餘月
        - C 欄：最新科餘月

        若 B/C 與目前輸入不符 → 直接 raise，避免刪錯批次。
        """
        subjects = []
        mismatch_rows = []

        for row in ws_summary.iter_rows(min_row=2):
            self._check_cancel()

            code = (str(row[0].value).strip() if row[0].value else "")
            make = (str(row[1].value).strip() if row[1].value else "")
            latest = (str(row[2].value).strip() if row[2].value else "")

            if not code:
                continue  # 空白列略過

            # 檢查年月是否一致
            if make != make_month or latest != latest_month:
                mismatch_rows.append((row[0].row, code, make, latest))
                continue

            subjects.append(code)

        if mismatch_rows:
            lines = [
                f"第 {r} 列：{code}（製作科餘月={make}，最新科餘月={latest}）"
                for (r, code, make, latest) in mismatch_rows
            ]
            detail = "\n".join(lines)
            raise Exception(
                "更新清單中的製作科餘月 / 最新科餘月與目前輸入不一致，"
                "請確認後再執行刪除。\n\n" + detail
            )

        self._log(f"📌 更新清單中共有 {len(subjects)} 個科目需要檢查。")
        return subjects

    # ---------- Step 2：處理單一科目分頁 ----------

    def _process_subject_sheet(self, subject_code: str):
        """
        對單一科目分頁執行：
        - 以摘要（E 欄）分組
        - 將該摘要底下的 F 欄、G 欄金額各自加總
        - 若一組摘要中 F 總額 == G 總額（誤差容許 0.001）→ 刪除該摘要下所有列
        回傳：刪除列數；若分頁不存在則回傳 None
        """
        if subject_code not in self.wb.sheetnames:
            self._log(f"⚠️ 找不到分頁「{subject_code}」，已略過。")
            return None

        ws = self.wb[subject_code]
        self._log(f"🔎 開始檢查分頁：{subject_code}")

        # 摘要 → { "rows": [index...], "sum_f": float, "sum_g": float }
        groups = defaultdict(lambda: {"rows": [], "sum_f": 0.0, "sum_g": 0.0})

        # 1️⃣ 先掃描所有列，建立分組
        for r in range(2, ws.max_row + 1):
            self._check_cancel()

            remark = ws[f"E{r}"].value
            if remark is None or str(remark).strip() == "":
                continue  # 沒摘要就不參與刪除判斷

            key = str(remark).strip()

            f_val = ws[f"F{r}"].value
            g_val = ws[f"G{r}"].value

            try:
                f_num = float(f_val) if f_val not in (None, "") else 0.0
            except Exception:
                f_num = 0.0

            try:
                g_num = float(g_val) if g_val not in (None, "") else 0.0
            except Exception:
                g_num = 0.0

            groups[key]["rows"].append(r)
            groups[key]["sum_f"] += f_num
            groups[key]["sum_g"] += g_num

        # 2️⃣ 判斷要刪除的列（F 總額 == G 總額 的摘要群組）
        rows_to_delete = []

        for remark, info in groups.items():
            self._check_cancel()

            sum_f = info["sum_f"]
            sum_g = info["sum_g"]

            if abs(sum_f - sum_g) < 0.001 and info["rows"]:
                rows_to_delete.extend(info["rows"])
                self._log(
                    f"🗑️ 摘要「{remark}」：F 合計={sum_f}, G 合計={sum_g} → 標記刪除 {len(info['rows'])} 列"
                )

        if not rows_to_delete:
            self._log(f"ℹ️ 分頁「{subject_code}」沒有符合刪除條件的明細。")
            return 0

        # 3️⃣ 由下往上刪除列，避免 row index 亂掉
        rows_to_delete = sorted(set(rows_to_delete), reverse=True)
        for r in rows_to_delete:
            self._check_cancel()
            ws.delete_rows(r, 1)

        return len(rows_to_delete)
