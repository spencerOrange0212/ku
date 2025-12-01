import os
import glob
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Tuple, Optional, Any, Dict


class SubjectPasteService:
    """
    負責「報表貼入科目」功能的業務邏輯服務。
    遵循 DI 原則，將所有檔案尋找、驗證、貼入邏輯封裝在此。
    """

    def __init__(self, logger=print, app=None):
        """
        初始化服務，儲存 Logger 和 App 實例 (依賴注入)
        """
        self.logger = logger
        self.app = app

    def _get_month_str(self, make_month: str) -> str:
        """解析月份字串，例如 '11401' -> '01月'"""
        if not make_month or len(make_month) < 2:
            return ""
        return make_month[-2:] + "月"

    # ==========================================
    # 1. 檔案搜尋與檢核工具
    # ==========================================

    def find_module_file(self, input_folder: str, make_month: str, vendor_id: str, module_name: str) -> str:
        """通用檔案搜尋器 (含 ID 備援與唯一性檢查)"""
        if not input_folder or not os.path.exists(input_folder):
            raise FileNotFoundError(f"輸入資料夾不存在：{input_folder}")

        month_str = self._get_month_str(make_month)
        month_folder = os.path.join(input_folder, month_str)

        if not os.path.exists(month_folder):
            raise FileNotFoundError(f"找不到月份資料夾：{month_folder}\n搜尋路徑：{month_folder}")

        # 搜尋規則：優先找 {module_name}* (無 ID)
        pattern = os.path.join(month_folder, f"{module_name}*")
        all_candidates = glob.glob(pattern)

        # 過濾邏輯 (只接受完全一致 或 帶有 _ 後綴的)
        target_stem = module_name
        valid_files = []

        for path in all_candidates:
            filename = os.path.basename(path)
            if filename.startswith("~$"): continue

            name_stem, _ = os.path.splitext(filename)
            # 完全一致 或 接底線
            if name_stem == target_stem or name_stem.startswith(target_stem + "_"):
                valid_files.append(path)

        # 備援搜尋 (找帶 vendor_id 的)
        if not valid_files:
            fallback = glob.glob(os.path.join(month_folder, f"{vendor_id}_{module_name}*"))
            valid_files = [f for f in fallback if not os.path.basename(f).startswith("~$")]

        if not valid_files:
            raise FileNotFoundError(f"❌ 找不到模組檔案：[{module_name}]")

        if len(valid_files) > 1:
            names = "\n".join([os.path.basename(f) for f in valid_files])
            raise ValueError(f"❌ 錯誤：模組 [{module_name}] 找到多個檔案，請保留唯一一個。\n{names}")

        return valid_files[0]

    def check_ledger_date_limit(self, file_path: str, make_month: str):
        """分類帳專用的日期檢查"""
        self.logger(f"正在檢查分類帳日期：{os.path.basename(file_path)}")
        try:
            target_year = int(make_month[:3])
            target_month = int(make_month[3:])
            df = pd.read_excel(file_path, usecols="A", dtype=str)
        except Exception as e:
            raise ValueError(f"無法讀取分類帳日期：{e}")

        error_list = []
        for row in df.itertuples():
            date_str = str(row[1]).strip()
            match = re.match(r"(\d{3})-(\d{1,2})-(\d{1,2})", date_str)
            if not match: continue

            y, m = int(match.group(1)), int(match.group(2))
            if (y > target_year) or (y == target_year and m > target_month):
                error_list.append(f"行 {row.Index + 2}: {date_str}")

        if error_list:
            msg = "\n".join(error_list[:5])
            raise ValueError(f"❌ 分類帳檢核失敗！發現未來日期：\n{msg}...")

        self.logger("✅ 日期檢核通過")

    # ==========================================
    # 2. 兩階段執行入口 (Orchestrator)
    # ==========================================

    def _validate_all_sources(self, input_folder: str, make_month: str, vendor_id: str,
                              required_tasks: List[Dict[str, Any]]):
        """
        階段一：檢查所有必要的來源檔案是否存在，並執行分類帳的內容驗證。
        """
        self.logger("🔍 開始進行貼入前的【所有檔案與內容】完整性檢查...")
        missing_files = []

        for config in required_tasks:
            module_name = config['module']  # ⭐️ 從字典中取值

            try:
                file_path = self.find_module_file(input_folder, make_month, vendor_id, module_name)

                # 關鍵邏輯：執行分類帳日期檢查 (只有在 check flag 存在時)
                if config['check'] == "LEDGER_DATE":
                    self.check_ledger_date_limit(file_path, make_month)

            except (FileNotFoundError, ValueError, RuntimeError) as e:
                # 捕捉到檔案找不到 OR 分類帳日期錯誤
                missing_files.append(str(e))
                continue  # 繼續檢查下一個檔案

        if missing_files:
            error_msg = "\n".join(missing_files)
            raise RuntimeError(
                f"❌ 貼入作業中止：以下 {len(missing_files)} 個必要檢查未通過，請補齊後再執行。\n{error_msg}"
            )

        self.logger("✅ 檔案與內容完整性檢查通過。")

    def execute_paste_task(self, input_folder: str, make_month: str, vendor_id: str, master_file_path: str):
        """
        主程式：執行三階段貼入作業 (檔案檢查 -> 分頁檢查 -> 執行)
        """
        # 1. 定義最終標準配置表
        REQUIRED_CONFIGS = [
            # 1. 資產負債表 (A:F, 貼入 A1)
            {"module": "資產負債表", "sheet": "資產負債表", "src_col_end": 6, "dest_row_start": 1, "dest_col_start": 1,
             "check": None},

            # 2. 綜合損益表 (A:G, 貼入 A1)
            {"module": "綜合損益表", "sheet": "綜合損益表", "src_col_end": 7, "dest_row_start": 1, "dest_col_start": 1,
             "check": None},

            # 3. 分類帳 (全貼, 貼入 A1, 需檢查日期)
            {"module": "分類帳", "sheet": "分類帳", "src_col_end": None, "dest_row_start": 1, "dest_col_start": 1,
             "check": "LEDGER_DATE"},

            # 4. 財產目錄 (全貼, 貼入 A1)
            {"module": "財產目錄", "sheet": "財產目錄", "src_col_end": None, "dest_row_start": 1, "dest_col_start": 1,
             "check": None},

            # 5. 綜合損益期別表 (動態裁剪末兩欄, 貼入 A1)
            {"module": "綜合損益期別表", "sheet": "綜合損益表-月份比較", "src_col_end": "DYNAMIC_CROP_2",
             "dest_row_start": 1, "dest_col_start": 1, "check": None},

            # ⭐️ 新增任務：綜合損益表邊欄 (末兩欄, 貼入 Z1) ⭐️
            {"module": "綜合損益期別表", "sheet": "綜合損益表-月份比較", "src_col_end": "SIDE_CROP_2", "dest_row_start": 1,
             "dest_col_start": 26, "check": None},

            # ⭐️ 6. 新任務：期別表負向索引邊欄 (貼入 AD/AE 欄) ⭐️
            {"module": "綜合損益期別表", "sheet": "綜合損益表-月份比較", "src_indices": [-6, -4], "dest_row_start": 1,
             "dest_col_start": 30, "check": None},
        ]

        # 2. 階段一：批次驗證檔案 (如果失敗，立即停止)
        self._validate_all_sources(input_folder, make_month, vendor_id, REQUIRED_CONFIGS)

        # 3. 階段二：開啟檔案與分頁檢查
        if not os.path.exists(master_file_path):
            raise FileNotFoundError(f"找不到科餘主檔：{master_file_path}")

        self.logger(f"📂 開始開啟科餘檔：{os.path.basename(master_file_path)} ...")

        try:
            wb = load_workbook(master_file_path)

            # ⭐️ 關鍵步驟：分頁預檢 ⭐️
            self._check_all_destination_sheets(wb, REQUIRED_CONFIGS)

            # 4. 階段三：執行貼入 (分頁已被確認存在，保證貼入不會失敗於找不到分頁)
            for config in REQUIRED_CONFIGS:
                # ⭐️ 檢查特殊處理邏輯 ⭐️
                if config.get("src_indices") is not None:
                    # 呼叫負向索引貼入專用方法
                    self._process_comparative_side_data(wb, input_folder, make_month, vendor_id, config)
                else:
                    # 其他所有標準和動態裁剪任務
                    self._process_task_unit(
                        wb, input_folder, make_month, vendor_id, config
                    )

            # 5. 存檔
            self.logger("💾 正在儲存檔案...")
            wb.save(master_file_path)
            self.logger("✅ 所有報表貼入作業完成！")

        except Exception as e:
            # 捕捉載入失敗、分頁檢查失敗、或執行時的錯誤
            raise RuntimeError(f"執行錯誤，已取消存檔：{e}")

    # ==========================================
    # 3. 核心統一執行邏輯 (單一任務處理器)
    # ==========================================

    def _process_task_unit(self, wb, input_folder, make_month, vendor_id, config: Dict[str, Any]):
        """
        通用流程：根據配置字典處理單一模組的所有 Find/Check/Paste 邏輯。
        """
        module_name = config['module']
        sheet_name = config['sheet']
        src_col_end = config['src_col_end']
        dest_row_start = config['dest_row_start']
        dest_col_start = config['dest_col_start']

        # 1. 找檔案 (略)
        try:
            file_path = self.find_module_file(input_folder, make_month, vendor_id, module_name)
        except FileNotFoundError:
            self.logger(f"   ⚠️ 警告：檔案 [{module_name}] 消失或無法讀取，跳過此模組。")
            return

        # 2. 讀取與裁剪
        try:
            # 讀取邏輯：使用 header=None 讀取所有數據，稍後手動切片
            df = pd.read_excel(file_path, header=None)
        except Exception as e:
            raise ValueError(f"讀取 {module_name} 失敗：{e}")

            # 根據配置執行裁剪
        if src_col_end == "DYNAMIC_CROP_2":
            # ⭐️ 修正：不再切除第一列，只裁剪末兩欄 ⭐️

            if df.shape[1] < 3:
                raise ValueError(f"[{module_name}] 欄位不足，無法裁剪末兩欄。")

            # 這是唯一需要的邏輯：保留所有行，只排除最後兩欄
            df_final = df.iloc[:, :-2]

        elif src_col_end == "SIDE_CROP_2":
            # 綜合損益表邊欄：只取末兩欄
            if df.shape[1] < 2:
                raise ValueError(f"[{module_name}] 欄位不足，無法複製末兩欄。")

            df_final = df.iloc[:, -2:]  # 裁剪：只保留末兩欄

        elif isinstance(src_col_end, int):
            # 標準報表的邏輯 (貼入 A1)：讀取 header=0，然後切前 N 欄
            df_final = df.iloc[:, :src_col_end]

            # 修正：由於標準報表是貼 A1，需要重新讀取確保 header=0
            # 避免 header=None 污染標準報表邏輯
            df_final = pd.read_excel(file_path, header=0).iloc[:, :src_col_end]

        else:
            # 分類帳/財產目錄的邏輯：全貼
            df_final = df

        # 3. 執行貼上
        self._write_sheet_data_from_df(
            wb,
            df_final,
            sheet_name,
            dest_row_start=dest_row_start,
            dest_col_start=dest_col_start,
            max_col_limit=src_col_end if isinstance(src_col_end, int) else None
        )

    def _process_comparative_side_data(self, wb, input_folder, make_month, vendor_id, config: Dict[str, Any]):
        """
        專門處理綜合損益期別表的負向索引邊欄數據貼入 (AD/AE)。
        修正：使用 List 串接法，穩定地將表頭作為數據第一行寫入。
        """
        module_name = config['module']
        sheet_name = config['sheet']

        # 1. 找檔案 (Find file)
        try:
            file_path = self.find_module_file(input_folder, make_month, vendor_id, module_name)
        except FileNotFoundError:
            self.logger(f"   ⚠️ 警告：邊欄貼入跳過，找不到 [{module_name}] 來源檔案。")
            return

        # 2. 讀取與裁剪 (Read and Crop Last Two Columns)
        try:
            # 使用 header=0 讀取，將 Row 1 轉為 Column Names
            df = pd.read_excel(file_path, header=0)

            # 裁剪：選取特定的負向索引欄位
            df_side_data = df.iloc[:, config['src_indices']]

            # ⭐️ 關鍵修正：使用 List 串接，將表頭注入數據體 ⭐️

            # 1. 獲取表頭列表 (Excel Row 1 的內容)
            header_list = [df_side_data.columns.tolist()]

            # 2. 獲取數據主體 (Excel Row 2 onwards)
            data_list = df_side_data.values.tolist()

            # 3. 合併：將表頭作為數據的第一行
            full_list = header_list + data_list

            # 4. 創建最終 DataFrame (欄位名稱為 0, 1, 2... 但內容是正確的)
            df_final = pd.DataFrame(full_list)

        except Exception as e:
            raise ValueError(f"讀取 {module_name} 邊欄失敗：{e}")

        # 3. 執行貼入 (使用統一寫入器)
        self._write_sheet_data_from_df(
            wb,
            df_final,  # ⭐️ 使用已包含表頭的 df_final ⭐️
            sheet_name,
            dest_row_start=config['dest_row_start'],
            dest_col_start=config['dest_col_start'],
            max_col_limit=None
        )
    def _write_sheet_data_from_df(self, wb, df_source, sheet_name, dest_row_start, dest_col_start, max_col_limit=None):
        """
        底層寫入邏輯：處理清除、位移寫入 (基於已裁剪的 DataFrame)。
        """

        # 1. 獲取工作表 (分頁檢查已在 Phase 2 完成)
        target_name_normalized = "".join(sheet_name.split())
        original_sheet_name = None
        for name in wb.sheetnames:
            if "".join(name.split()) == target_name_normalized:
                original_sheet_name = name
                break

        ws = wb[original_sheet_name]

        # 2. 清除舊資料
        paste_width = df_source.shape[1]
        current_max_row = ws.max_row

        # 決定清除的寬度：從 dest_col_start 開始，到貼入數據的寬度
        max_col_to_clear = dest_col_start + paste_width - 1

        # 遵守欄位限制 (例如資產負債表的 A:F 限制)
        if max_col_limit is not None and max_col_to_clear > max_col_limit:
            max_col_to_clear = max_col_limit

        # 清除範圍：從起始列開始，清除到最大列，起始欄位到結束欄位
        if current_max_row >= dest_row_start:
            for row in ws.iter_rows(min_row=dest_row_start, max_row=current_max_row,
                                    min_col=dest_col_start, max_col=max_col_to_clear):
                for cell in row:
                    cell.value = None

        # 3. 寫入新資料
        # header=False 因為我們已經在處理器中決定了是否要讀取表頭 (Row 1 vs Row 2)
        rows = dataframe_to_rows(df_source, index=False, header=False)
        end_col_ws = dest_col_start

        for r_idx, row in enumerate(rows, dest_row_start):  # r_idx 從 dest_row_start 開始
            for c_idx_data, value in enumerate(row, 0):

                col_index_ws = c_idx_data + dest_col_start  # 實際寫入的 Excel 欄位索引

                # 遵守欄位限制 (例如資產負債表的 A:F 限制)
                if max_col_limit is not None and col_index_ws > max_col_limit:
                    break

                ws.cell(row=r_idx, column=col_index_ws, value=value)
                end_col_ws = col_index_ws

        # 4. Log 訊息
        end_col_letter = chr(ord('A') + end_col_ws - 1)
        self.logger(
            f"      ✅ 已更新 {len(df_source)} 筆資料 (範圍: {chr(ord('A') + dest_col_start - 1)}{dest_row_start}~{end_col_letter}{r_idx - 1})")

    def _check_all_destination_sheets(self, wb, required_tasks: List[Dict[str, Any]]):
        """
        階段二：檢查目標工作簿中所有分頁名稱是否都存在。
        """
        self.logger("🔍 開始進行目標分頁名稱檢查...")
        missing_sheets = []

        for config in required_tasks:
            sheet_name = config['sheet']  # ⭐️ 依賴字典結構 ⭐️

            # 複製 _paste_data_to_sheet 裡面的檢查邏輯
            target_name_normalized = "".join(sheet_name.split())
            original_sheet_name = None

            for name in wb.sheetnames:
                sheet_name_normalized = "".join(name.split())
                if sheet_name_normalized == target_name_normalized:
                    original_sheet_name = name
                    break

            if not original_sheet_name:
                # 如果找不到，記錄錯誤
                missing_sheets.append(f"分頁 [{sheet_name}]")

        if missing_sheets:
            error_summary = "\n".join(missing_sheets)
            raise RuntimeError(
                f"❌ 嚴重錯誤：目標檔案中缺少以下分頁，貼入中止：\n{error_summary}"
            )

        self.logger("✅ 目標分頁名稱檢查通過。")
